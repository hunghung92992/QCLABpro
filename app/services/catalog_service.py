# -*- coding: utf-8 -*-
"""
app/services/catalog_service.py
(SAFE VERSION: Handles missing 'level' column gracefully + Auto Migration)
"""

from __future__ import annotations
from typing import Optional, List, Any, Dict
import json
import datetime as dt

# SQLAlchemy Imports
from sqlalchemy import text, func, or_, desc
from sqlalchemy.orm import Session
from sqlalchemy.exc import OperationalError

# App Imports
from app.core.database_orm import SessionLocal
from app.models.catalog_models import CatalogLot, CatalogAnalyte
from app.models.core_models import Department
from app.models.iqc_models import IQCResult
from app.services.iqc_service import IQCService
from app.utils.validators import to_float_safe as _to_float
from app.utils import analytics
import datetime


# ------------------------------------------------------------
# Helpers
# ------------------------------------------------------------
def _parse_meta_from_note(note: Optional[str]) -> Dict[str, Any]:
    """Parse JSON note to extract metadata."""
    if not note:
        return {}
    if not str(note).strip().startswith("{"):
        return {"note_text": str(note)}
    try:
        return json.loads(note)
    except Exception:
        return {}


def assert_non_empty(s: Optional[str], field: str):
    """Validate that string is not empty."""
    if s is None or not str(s).strip():
        raise ValueError(f"{field} không được trống")


# ------------------------------------------------------------
# CatalogService
# ------------------------------------------------------------
class CatalogService:

    def __init__(self):
        # 🌟 Sử dụng SessionLocal an toàn cho Thread
        self.db: Session = SessionLocal()

        # 🌟 TỰ ĐỘNG BƠM CỘT THIẾU (Giải quyết Phase 3.1 & 3.3)
        self._ensure_schema_columns()

        self._has_level_column = True
        self._has_note_column = True

    def __del__(self):
        try:
            self.db.close()
        except:
            pass

    def _ensure_schema_columns(self):
        """Tự động kiểm tra và vá lỗi cấu trúc Database (Migration ngầm)"""
        migrations = [
            ("catalog_lots", "level", "VARCHAR(50)"),
            ("catalog_lots", "device_sample_id", "VARCHAR(100)"),
            ("catalog_analytes", "level", "VARCHAR(50)"),
            ("catalog_analytes", "note", "TEXT"),
            ("catalog_analytes", "lims_code", "VARCHAR(100)"),
            ("iqc_results", "level", "VARCHAR(50)"),  # Xử lý Phase 3.1
            ("iqc_results", "note", "TEXT")  # Xử lý Phase 3.1
        ]

        for table, column, col_type in migrations:
            try:
                # Kiểm tra thử bằng cách SELECT
                self.db.execute(text(f"SELECT {column} FROM {table} LIMIT 1"))
            except Exception:
                self.db.rollback()
                try:
                    # Bơm cột nếu thiếu
                    self.db.execute(text(f"ALTER TABLE {table} ADD COLUMN {column} {col_type}"))
                    self.db.commit()
                    print(f"🔧 [Migration] Đã vá thành công cột '{column}' cho bảng '{table}'")
                except Exception as e:
                    self.db.rollback()
                    # Cột có thể đã tồn tại hoặc bảng chưa được tạo
                    pass

    def _get_db(self):
        return SessionLocal()

    def _get_or_create_dept_id(self, dept_name: str) -> Optional[str]:
        """Helper: Tìm ID phòng ban từ tên. TỰ ĐỘNG TẠO MỚI nếu chưa có."""
        if not dept_name or not dept_name.strip():
            return None

        name_clean = dept_name.strip()

        dept = self.db.query(Department).filter(
            func.lower(Department.name) == name_clean.lower()
        ).first()

        if dept: return dept.id

        try:
            new_dept = Department(name=name_clean, active=1, sync_flag=1)
            self.db.add(new_dept)
            self.db.commit()
            return new_dept.id
        except Exception as e:
            self.db.rollback()
            print(f"Error auto-creating department '{name_clean}': {e}")
            return None

    def _get_dept_name_by_id(self, dept_id: str) -> str:
        if not dept_id: return ""
        dept = self.db.query(Department).filter(Department.id == str(dept_id)).first()
        return dept.name if dept else ""

    # ----------------------------------------------------------------
    # 1. LOGIC STATS
    # ----------------------------------------------------------------
    def calculate_lot_stats(self, department: str, test_name: str, level: str, lot_no: str) -> Dict[str, Any]:
        """Tính toán lại Mean, SD, CV% từ dữ liệu IQC."""
        MIN_N = 20
        iqc_service = IQCService()
        history = iqc_service.get_history(
            department=department, test_code=test_name, level=level,
            lot_no=lot_no, limit=500, active_only=True, sort_order="DESC"
        )
        values = []
        for r in history:
            if r.get('value_num') is not None and r.get('pass_fail') == 1:
                val = _to_float(r.get('value_num'))
                if val is not None:
                    values.append(val)

        n = len(values)
        if n < MIN_N:
            return {
                "n": n, "mean": None, "sd": None, "cv": None,
                "error": f"Không đủ điểm đạt chuẩn (N={n}, Yêu cầu ≥{MIN_N} điểm Đạt)."
            }
        stats = analytics.compute_stats(values)
        stats["n"] = n
        return stats

    # ----------------------------------------------------------------
    # 2. CATALOG & META
    # ----------------------------------------------------------------
    def list_tests_by_department(self, department: str) -> list[str]:
        dep = (department or "").strip().lower()
        if not dep: return []
        out: set[str] = set()
        try:
            # 1. Legacy Catalog
            try:
                sql = text("SELECT test_name FROM lab_test_catalog WHERE LOWER(department) = :d AND is_active=1")
                res = self.db.execute(sql, {"d": dep}).fetchall()
                for r in res: out.add(r[0])
            except Exception:
                pass

            # 2. CatalogAnalyte
            q = self.db.query(CatalogAnalyte.test_name) \
                .join(CatalogLot, CatalogAnalyte.lot_id == CatalogLot.id) \
                .join(Department, CatalogLot.department_id == Department.id) \
                .filter(func.lower(Department.name) == dep, CatalogAnalyte.sync_flag != 2)
            for r in q.distinct().all():
                out.add(r.test_name)

            # 3. Fallback: IQCResult
            try:
                q_iqc = self.db.query(IQCResult.test_code) \
                    .join(Department, IQCResult.department_id == Department.id) \
                    .filter(func.lower(Department.name) == dep)
                for r in q_iqc.distinct().all():
                    if r.test_code: out.add(r.test_code)
            except Exception:
                pass
        except Exception as e:
            print(f"[CatalogService] List tests error: {e}")

        return sorted(list(out), key=str.lower)

    def get_test_meta(self, department: str, test_name: str) -> dict:
        dep = (department or "").strip().lower()
        t = (test_name or "").strip().lower()
        try:
            sql = text(
                "SELECT data_type, default_unit FROM lab_test_catalog WHERE LOWER(department) = :d AND LOWER(test_name) = :t AND is_active=1")
            row = self.db.execute(sql, {"d": dep, "t": t}).first()
            if row: return {"data_type": row[0], "unit": row[1]}

            analyte = self.db.query(CatalogAnalyte) \
                .join(CatalogLot, CatalogAnalyte.lot_id == CatalogLot.id) \
                .join(Department, CatalogLot.department_id == Department.id) \
                .filter(func.lower(Department.name) == dep, func.lower(CatalogAnalyte.test_name) == t) \
                .order_by(desc(CatalogAnalyte.created_at)).first()
            if analyte:
                meta = _parse_meta_from_note(analyte.note)
                dtype = meta.get("data_type", "Quant")
                return {"data_type": dtype, "unit": analyte.unit}
        except Exception:
            pass
        return {"data_type": None, "unit": None}

    def upsert_catalog(self, department: str, test_name: str, *, data_type: str | None, default_unit: str | None):
        if not department or not test_name: return
        try:
            check_sql = text(
                "SELECT department FROM lab_test_catalog WHERE LOWER(department)=:d AND LOWER(test_name)=:t")
            existing = self.db.execute(check_sql, {"d": department.lower(), "t": test_name.lower()}).first()
            if existing:
                upd = text(
                    "UPDATE lab_test_catalog SET data_type=:dt, default_unit=:u, is_active=1 WHERE LOWER(department)=:d AND LOWER(test_name)=:t")
                self.db.execute(upd,
                                {"dt": data_type, "u": default_unit, "d": department.lower(), "t": test_name.lower()})
            else:
                ins = text(
                    "INSERT INTO lab_test_catalog (department, test_name, data_type, default_unit, is_active) VALUES (:d, :t, :dt, :u, 1)")
                self.db.execute(ins, {"d": department, "t": test_name, "dt": data_type, "u": default_unit})
            self.db.commit()
        except Exception:
            self.db.rollback()

    # ----------------------------------------------------------------
    # 3. LOT CRUD
    # ----------------------------------------------------------------
    def _lot_to_dict(self, lot: CatalogLot, dept_name: str = None) -> Dict:
        if dept_name is None:
            dept_name = self._get_dept_name_by_id(lot.department_id) if lot.department_id else ""

        return {
            "id": str(lot.id),
            "name": lot.lot_name,
            "lot_name": lot.lot_name,
            "lot": lot.lot_code,
            "lot_no": lot.lot_code,
            "mfg_date": lot.mfg_date,
            "mfg": lot.mfg_date,
            "exp_date": lot.exp_date,
            "exp": lot.exp_date,
            "expiry_date": lot.exp_date,
            "department": dept_name,
            "status": lot.status,
            "level": lot.level,
            "active": 1 if lot.status == 'active' else 0,
            "device_sample_id": getattr(lot, 'device_sample_id', '')
        }

    def exists_lot(self, department: Optional[str], lot_code: str, *, exclude_id: Optional[str] = None) -> bool:
        assert_non_empty(lot_code, "lot")
        try:
            q = self.db.query(CatalogLot).filter(func.lower(CatalogLot.lot_code) == lot_code.strip().lower(),
                                                 CatalogLot.sync_flag != 2)
            if department:
                dept_id = self._get_or_create_dept_id(department)
                if dept_id: q = q.filter(CatalogLot.department_id == dept_id)
            if exclude_id: q = q.filter(CatalogLot.id != str(exclude_id))
            return q.first() is not None
        except Exception:
            return False

    def list_lots(self, active_only: bool = False) -> List[Dict]:
        try:
            q = self.db.query(CatalogLot, Department.name).outerjoin(Department,
                                                                     CatalogLot.department_id == Department.id).filter(
                CatalogLot.sync_flag != 2)
            if active_only: q = q.filter(CatalogLot.status == 'active')
            q = q.order_by(desc(CatalogLot.created_at))
            return [self._lot_to_dict(l, d_name) for l, d_name in q.all()]
        except Exception as e:
            print(f"Error list_lots: {e}")
            return []

    def get_lot(self, lot_id: str) -> Dict[str, Any]:
        try:
            res = self.db.query(CatalogLot, Department.name).outerjoin(Department,
                                                                       CatalogLot.department_id == Department.id).filter(
                CatalogLot.id == str(lot_id)).first()
            if res: return self._lot_to_dict(res[0], res[1])
            return {}
        except Exception:
            return {}

    def create_lot(self, name: str, lot: str, mfg_date: Optional[str], exp_date: Optional[str],
                   department: Optional[str] = None, status: Optional[str] = "active",
                   level: Optional[str] = None, device_sample_id: Optional[str] = None) -> str:
        assert_non_empty(lot, "lot")
        dept_id = self._get_or_create_dept_id(department)

        if department and dept_id:
            existing = self.db.query(CatalogLot).filter(
                func.lower(CatalogLot.lot_code) == lot.strip().lower(),
                CatalogLot.department_id == dept_id,
                CatalogLot.sync_flag != 2
            ).first()
            if existing: raise ValueError("Trùng LOT trong cùng phòng ban")

        try:
            new_lot = CatalogLot(
                lot_name=name, lot_code=lot.strip(), mfg_date=mfg_date, exp_date=exp_date,
                department_id=dept_id, status=status, level=level,
                device_sample_id=device_sample_id, sync_flag=1, created_at=datetime.datetime.utcnow()
            )
            self.db.add(new_lot)
            self.db.commit()
            return str(new_lot.id)
        except Exception as e:
            self.db.rollback()
            raise e

    def update_lot(self, lot_id: str, **kwargs) -> bool:
        try:
            l_obj = self.db.query(CatalogLot).get(str(lot_id))
            if not l_obj: return False

            if 'department' in kwargs: l_obj.department_id = self._get_or_create_dept_id(kwargs['department'])
            if 'name' in kwargs: l_obj.lot_name = kwargs['name']
            if 'lot_name' in kwargs: l_obj.lot_name = kwargs['lot_name']
            if 'lot' in kwargs: l_obj.lot_code = kwargs['lot'].strip()

            if 'mfg_date' in kwargs:
                l_obj.mfg_date = kwargs['mfg_date']
            elif 'mfg' in kwargs:
                l_obj.mfg_date = kwargs['mfg']

            if 'exp_date' in kwargs:
                l_obj.exp_date = kwargs['exp_date']
            elif 'exp' in kwargs:
                l_obj.exp_date = kwargs['exp']

            if 'status' in kwargs: l_obj.status = kwargs['status']
            if 'level' in kwargs: l_obj.level = kwargs['level']
            if 'device_sample_id' in kwargs: l_obj.device_sample_id = kwargs['device_sample_id']

            l_obj.sync_flag = 1
            self.db.commit()
            return True
        except Exception:
            self.db.rollback()
            return False

    def delete_lot(self, lot_id: str) -> bool:
        try:
            l_obj = self.db.query(CatalogLot).get(str(lot_id))
            if l_obj:
                l_obj.sync_flag = 2
                self.db.query(CatalogAnalyte).filter(CatalogAnalyte.lot_id == str(lot_id)).update(
                    {CatalogAnalyte.sync_flag: 2})
                self.db.commit()
                return True
            return False
        except Exception:
            self.db.rollback()
            return False

    def delete_lots(self, lot_ids: List[str]) -> int:
        return sum(1 for i in lot_ids if self.delete_lot(i))

    def clone_details(self, src_lot_id: str, dst_lot_id: str, *, overwrite: bool = False) -> int:
        try:
            if overwrite:
                self.db.query(CatalogAnalyte).filter(CatalogAnalyte.lot_id == str(dst_lot_id)).update(
                    {CatalogAnalyte.sync_flag: 2}, synchronize_session=False)

            srcs = self.db.query(CatalogAnalyte).filter(CatalogAnalyte.lot_id == str(src_lot_id),
                                                        CatalogAnalyte.sync_flag != 2).all()
            cnt = 0
            for r in srcs:
                params = {
                    'lot_id': str(dst_lot_id),
                    'department': getattr(r, 'department', None),
                    'test_name': r.test_name,
                    'test_code': getattr(r, 'test_code', None),
                    'mean': getattr(r, 'mean', 0.0),
                    'sd': getattr(r, 'sd', 0.0),
                    'tea': getattr(r, 'tea', 0.0),
                    'unit': r.unit,
                    'category': getattr(r, 'category', None),
                    'data_type': getattr(r, 'data_type', 'Quant'),
                    'reference_range': getattr(r, 'reference_range', None),
                    'westgard_rules': r.westgard_rules,
                    'sort_order': getattr(r, 'sort_order', 0),
                    'sync_flag': 1
                }
                if hasattr(r, 'note') and r.note: params['note'] = r.note
                if hasattr(r, 'level') and r.level: params['level'] = r.level
                if hasattr(r, 'lims_code'): params['lims_code'] = r.lims_code

                self.db.add(CatalogAnalyte(**params))
                cnt += 1
            self.db.commit()
            return cnt
        except Exception as e:
            self.db.rollback()
            print(f"❌ [CatalogService] Lỗi Clone Details: {e}")
            return 0

    # ----------------------------------------------------------------
    # 4. LOT DETAIL CRUD
    # ----------------------------------------------------------------
    def _analyte_to_dict(self, a: CatalogAnalyte, dept_name: str = "") -> Dict:
        note_content = getattr(a, 'note', None)
        meta = _parse_meta_from_note(note_content)
        level_value = getattr(a, 'level', None) or meta.get('level', '')

        return {
            "id": str(a.id),
            "lot_id": str(a.lot_id),
            "department": dept_name,
            "test_name": a.test_name,
            "test_code": a.test_code,
            "level": level_value,
            "mean": getattr(a, "mean", None),
            "sd": getattr(a, "sd", None),
            "tea": getattr(a, "tea", None),
            "lims_code": getattr(a, "lims_code", ""),
            "note": note_content,
            "active": 1,
            "data_type": meta.get("data_type", "Quant"),
            "unit": a.unit,
            "reference_range": meta.get("reference_range", ""),
            "category": meta.get("category", ""),
            "sort_order": a.sort_order,
            "westgard_rules": a.westgard_rules
        }

    def list_details(self, lot_id: str) -> List[Dict]:
        try:
            info = self.db.query(CatalogLot, Department.name).outerjoin(Department,
                                                                        CatalogLot.department_id == Department.id).filter(
                CatalogLot.id == str(lot_id)).first()
            d_name = info[1] if info else ""
            rows = self.db.query(CatalogAnalyte).filter(CatalogAnalyte.lot_id == str(lot_id),
                                                        CatalogAnalyte.sync_flag != 2).order_by(
                CatalogAnalyte.sort_order, CatalogAnalyte.test_name).all()
            return [self._analyte_to_dict(r, d_name) for r in rows]
        except Exception:
            return []

    def get_detail(self, detail_id: str) -> Dict[str, Any]:
        try:
            d = self.db.query(CatalogAnalyte).get(str(detail_id))
            if not d: return {}
            info = self.db.query(CatalogLot, Department.name).outerjoin(Department,
                                                                        CatalogLot.department_id == Department.id).filter(
                CatalogLot.id == d.lot_id).first()
            d_name = info[1] if info else ""
            return self._analyte_to_dict(d, d_name)
        except Exception:
            return {}

    def create_detail(self, lot_id: str, department: Optional[str], test_name: str,
                      mean: Optional[float], sd: Optional[float], tea: Optional[float],
                      note: Optional[str], level: Optional[str] = None,
                      data_type: Optional[str] = None, unit: Optional[str] = None,
                      reference_range: Optional[str] = None, category: Optional[str] = None,
                      sort_order: Optional[int] = 0) -> str:
        assert_non_empty(test_name, "test_name")
        meta = {}
        if note and str(note).strip().startswith("{"):
            try:
                meta = json.loads(note)
            except:
                meta["note_text"] = note
        elif note:
            meta["note_text"] = note

        if category: meta["category"] = category
        if data_type: meta["data_type"] = data_type
        if unit: meta["unit"] = unit
        if reference_range: meta["reference_range"] = reference_range
        final_note = json.dumps(meta, ensure_ascii=False) if meta else None

        try:
            params = {
                'lot_id': str(lot_id), 'test_name': test_name.strip(), 'test_code': test_name.strip(),
                'mean': mean, 'sd': sd, 'tea': tea, 'unit': unit, 'sort_order': sort_order,
                'sync_flag': 1, 'note': final_note, 'level': level
            }
            new_a = CatalogAnalyte(**params)
            self.db.add(new_a)
            self.db.commit()
            return str(new_a.id)
        except Exception as e:
            self.db.rollback()
            raise e

    def update_detail(self, detail_id: str, **kwargs) -> bool:
        try:
            d = self.db.query(CatalogAnalyte).get(str(detail_id))
            if not d: return False

            if 'test_name' in kwargs and kwargs['test_name']:
                d.test_name = kwargs['test_name'].strip()
                d.test_code = kwargs['test_name'].strip()

            if 'mean' in kwargs: d.mean = kwargs['mean']
            if 'sd' in kwargs: d.sd = kwargs['sd']
            if 'tea' in kwargs: d.tea = kwargs['tea']
            if 'unit' in kwargs: d.unit = kwargs['unit']
            if 'sort_order' in kwargs: d.sort_order = kwargs['sort_order']
            if 'note' in kwargs: d.note = kwargs['note']
            if 'level' in kwargs: d.level = kwargs['level']
            if 'lims_code' in kwargs: d.lims_code = kwargs['lims_code']

            d.sync_flag = 1
            self.db.commit()
            return True
        except Exception:
            self.db.rollback()
            return False

    def delete_detail(self, detail_id: str) -> bool:
        try:
            d = self.db.query(CatalogAnalyte).get(str(detail_id))
            if d:
                d.sync_flag = 2
                self.db.commit()
                return True
            return False
        except Exception:
            self.db.rollback()
            return False

    def delete_details(self, detail_ids: List[str]) -> int:
        return sum(1 for i in detail_ids if self.delete_detail(i))

    # ----------------------------------------------------------------
    # 5. SEARCH & UI HELPERS
    # ----------------------------------------------------------------
    def search_lots(self, department: Optional[str] = None, lot: Optional[str] = None,
                    status: Optional[str] = None, mfg_from: Optional[str] = None, mfg_to: Optional[str] = None,
                    exp_from: Optional[str] = None, exp_to: Optional[str] = None) -> List[Dict]:
        try:
            q = self.db.query(CatalogLot, Department.name).outerjoin(Department,
                                                                     CatalogLot.department_id == Department.id).filter(
                CatalogLot.sync_flag != 2)
            if department:
                d_id = self._get_or_create_dept_id(department)
                if d_id: q = q.filter(CatalogLot.department_id == d_id)
            if lot: q = q.filter(func.lower(CatalogLot.lot_code).contains(lot.lower()))
            if status: q = q.filter(func.lower(CatalogLot.status) == status.lower())
            if mfg_from: q = q.filter(CatalogLot.mfg_date >= mfg_from)
            if mfg_to: q = q.filter(CatalogLot.mfg_date <= mfg_to)
            if exp_from: q = q.filter(CatalogLot.exp_date >= exp_from)
            if exp_to: q = q.filter(CatalogLot.exp_date <= exp_to)

            q = q.order_by(desc(CatalogLot.created_at))
            return [self._lot_to_dict(l, d_name) for l, d_name in q.all()]
        except Exception:
            return []

    def get_target_by_lot(self, test_code: str, level_str: str, lot_no: Optional[str]) -> Optional[Dict[str, Any]]:
        if not test_code or not level_str: return None
        try:
            lot_id = None
            if lot_no:
                l = self.db.query(CatalogLot).filter(CatalogLot.lot_code == lot_no, CatalogLot.sync_flag != 2).first()
                if l: lot_id = l.id
            if lot_no and not lot_id: return None

            q = self.db.query(CatalogAnalyte).filter(func.lower(CatalogAnalyte.test_name) == test_code.lower(),
                                                     CatalogAnalyte.sync_flag != 2)
            if lot_id:
                q = q.filter(CatalogAnalyte.lot_id == lot_id)
            else:
                q = q.order_by(desc(CatalogAnalyte.updated_at))

            q = q.join(CatalogLot, CatalogAnalyte.lot_id == CatalogLot.id)
            lv_clean = level_str.replace("Level", "").replace("L", "").strip()
            q = q.filter(or_(CatalogLot.level == level_str, CatalogLot.level == lv_clean))

            r = q.first()
            if r:
                info = self.db.query(CatalogLot, Department.name).outerjoin(Department,
                                                                            CatalogLot.department_id == Department.id).filter(
                    CatalogLot.id == r.lot_id).first()
                d_name = info[1] if info else ""
                return self._analyte_to_dict(r, d_name)
            return None
        except Exception:
            return None

    def get_details_by_lot(self, lot_id: str) -> List[Dict[str, Any]]:
        return self.list_details(lot_id)

    def list_devices_by_department(self, department: str) -> List[Dict[str, str]]:
        dep_txt = (department or "").strip()
        if not dep_txt: return []
        try:
            dep_id = self._get_or_create_dept_id(dep_txt)
            if not dep_id: return []
            from app.models.core_models import Device
            devices = self.db.query(Device).filter(Device.department_id == str(dep_id)).all()
            results, seen = [], set()
            for d in devices:
                name = d.name or d.code or "No Name"
                if name not in seen:
                    results.append({"id": str(d.id), "name": name})
                    seen.add(name)
            return sorted(results, key=lambda x: x['name'])
        except Exception:
            return []

    def list_active_lots_by_level(self, department_name: str | int, *, only_valid_expiry: bool = False) -> Dict[
        str, List[Dict[str, str]]]:
        lots_by_level = {'L1': [], 'L2': [], 'L3': []}
        try:
            all_lots = self.db.query(CatalogLot).filter(CatalogLot.status == 'active',
                                                        CatalogLot.sync_flag != 2).order_by(
                desc(CatalogLot.created_at)).all()
            target_dept_name = str(department_name or "").strip().lower()

            for lot in all_lots:
                is_dept_match = True
                if target_dept_name and target_dept_name != "tất cả":
                    lot_dept_name = self._get_dept_name_by_id(
                        lot.department_id).strip().lower() if lot.department_id else ""
                    if lot.department_id and target_dept_name not in lot_dept_name and lot_dept_name not in target_dept_name:
                        is_dept_match = False
                if not is_dept_match: continue

                if only_valid_expiry and lot.exp_date:
                    try:
                        if lot.exp_date < dt.date.today().isoformat(): continue
                    except:
                        pass

                raw_level = str(lot.level or "").upper().strip()
                lot_code_upper = str(lot.lot_code).upper().strip()
                level_key = None

                if raw_level in ["1", "L1", "LEVEL 1", "LOW", "THẤP"]:
                    level_key = "L1"
                elif raw_level in ["2", "L2", "LEVEL 2", "NORMAL", "TB", "TRUNG"]:
                    level_key = "L2"
                elif raw_level in ["3", "L3", "LEVEL 3", "HIGH", "CAO"]:
                    level_key = "L3"

                if not level_key:
                    if "L1" in lot_code_upper or "LEVEL 1" in lot_code_upper or "LEV 1" in lot_code_upper:
                        level_key = "L1"
                    elif "L2" in lot_code_upper or "LEVEL 2" in lot_code_upper or "LEV 2" in lot_code_upper:
                        level_key = "L2"
                    elif "L3" in lot_code_upper or "LEVEL 3" in lot_code_upper or "LEV 3" in lot_code_upper:
                        level_key = "L3"

                if level_key:
                    item = {"lot_no": lot.lot_code, "expiry_date": lot.exp_date, "id": str(lot.id)}
                    if not any(x['lot_no'] == item['lot_no'] for x in lots_by_level[level_key]):
                        lots_by_level[level_key].append(item)

            return lots_by_level
        except Exception:
            return {'L1': [], 'L2': [], 'L3': []}

    # ----------------------------------------------------------------
    # 8. EXCEL & UTILS
    # ----------------------------------------------------------------
    def export_excel(self, filepath: str, lot_id: Optional[str] = None) -> None:
        try:
            from openpyxl import Workbook
        except ImportError:
            return
        wb = Workbook()
        ws1 = wb.active;
        ws1.title = "LOTS"
        ws2 = wb.create_sheet("ANALYTES")
        lots = self.list_lots()
        if lots:
            headers = list(lots[0].keys())
            ws1.append(headers)
            for r in lots: ws1.append([str(r.get(h)) for h in headers])

        q = self.db.query(CatalogAnalyte).filter(CatalogAnalyte.sync_flag != 2)
        if lot_id: q = q.filter(CatalogAnalyte.lot_id == str(lot_id))
        analytes = []
        for a in q.all():
            info = self.db.query(CatalogLot, Department.name).outerjoin(Department).filter(
                CatalogLot.id == a.lot_id).first()
            d = info[1] if info else ""
            analytes.append(self._analyte_to_dict(a, d))
        if analytes:
            headers = list(analytes[0].keys())
            ws2.append(headers)
            for r in analytes: ws2.append([str(r.get(h)) for h in headers])
        wb.save(filepath)

    def import_excel(self, filepath: str, default_lot_id: Optional[str] = None) -> Dict[str, int]:
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError("Yêu cầu thư viện 'openpyxl'")

        wb = load_workbook(filepath, data_only=True)
        stats = {"lots": 0, "analytes": 0}
        lot_map = {}

        if not default_lot_id and "LOTS" in wb.sheetnames:
            ws = wb["LOTS"]
            headers = [c.value for c in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                r = dict(zip(headers, row))
                lot_code = str(r.get("lot_no") or r.get("lot") or "")
                if not lot_code: continue

                dept_id = self._get_or_create_dept_id(r.get("department"))
                existing = self.db.query(CatalogLot).filter(CatalogLot.lot_code == lot_code,
                                                            CatalogLot.department_id == dept_id,
                                                            CatalogLot.sync_flag != 2).first()

                if existing:
                    if r.get("mfg_date"): existing.mfg_date = r.get("mfg_date")
                    if r.get("expiry_date"): existing.exp_date = r.get("expiry_date")
                    existing.sync_flag = 1
                    lot_map[lot_code] = existing.id
                else:
                    new_lot = CatalogLot(
                        lot_name=r.get("name") or r.get("lot_name"), lot_code=lot_code,
                        mfg_date=r.get("mfg_date"), exp_date=r.get("expiry_date"),
                        department_id=dept_id, status=r.get("status") or "active", sync_flag=1
                    )
                    self.db.add(new_lot)
                    self.db.commit()
                    lot_map[lot_code] = new_lot.id
                    stats["lots"] += 1

        sheet_name = "ANALYTES" if "ANALYTES" in wb.sheetnames else (
            wb.sheetnames[0] if len(wb.sheetnames) == 1 else None)
        if sheet_name:
            ws = wb[sheet_name]
            headers = [c.value for c in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                r = dict(zip(headers, row))
                test_name = r.get("test_name")
                if not test_name: continue

                target_lot_id = default_lot_id
                if not target_lot_id:
                    file_lot_id = str(r.get("lot_id") or "")
                    if file_lot_id and self.db.query(CatalogLot).get(file_lot_id): target_lot_id = file_lot_id
                if not target_lot_id: continue

                test_code = str(r.get("test_code") or test_name).strip()
                meta_import = {
                    "level": r.get("level"), "category": r.get("category"),
                    "data_type": r.get("data_type"), "unit": r.get("unit"), "reference_range": r.get("reference_range")
                }
                note_json = json.dumps(meta_import, ensure_ascii=False)

                existing_a = self.db.query(CatalogAnalyte).filter(CatalogAnalyte.lot_id == target_lot_id,
                                                                  CatalogAnalyte.test_code == test_code,
                                                                  CatalogAnalyte.sync_flag != 2).first()

                if existing_a:
                    if r.get("mean"): existing_a.mean = _to_float(r.get("mean"))
                    if r.get("sd"): existing_a.sd = _to_float(r.get("sd"))
                    if r.get("tea"): existing_a.tea = _to_float(r.get("tea"))
                    existing_a.note = note_json
                    existing_a.level = r.get("level")
                    existing_a.sync_flag = 1
                else:
                    new_a = CatalogAnalyte(
                        lot_id=target_lot_id, test_code=test_code, test_name=str(test_name).strip(),
                        mean=_to_float(r.get("mean")), sd=_to_float(r.get("sd")), tea=_to_float(r.get("tea")),
                        unit=r.get("unit"), note=note_json, sort_order=_to_float(r.get("sort_order")) or 0, sync_flag=1
                    )
                    new_a.level = r.get("level")
                    self.db.add(new_a)
                    stats["analytes"] += 1
                self.db.commit()

            return stats

    def update_analyte_mapping(self, detail_id: str, lims_code: str):
        try:
            analyte = self.db.query(CatalogAnalyte).filter(CatalogAnalyte.id == str(detail_id)).first()
            if analyte:
                analyte.lims_code = str(lims_code).strip() if lims_code else None
                analyte.sync_flag = 1
                self.db.commit()
                return True
            return False
        except Exception as e:
            self.db.rollback()
            print(f"❌ Mapping Error: {e}")
            return False