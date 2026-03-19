# -*- coding: utf-8 -*-
"""
app/services/iqc_service.py
(PHIÊN BẢN MASTER - TÍCH HỢP WESTGARD PIPELINE & ORM)
Tự động tính toán Z-Score, Pass/Fail và dán nhãn lỗi Westgard.
"""

from __future__ import annotations
from typing import List, Dict, Any, Optional, Tuple
import datetime as dt
import json, os

# Import SQLAlchemy
from sqlalchemy import desc, asc, func
# Imports từ core ORM
from app.core.database_orm import SessionLocal
from app.models.iqc_models import IQCRun, IQCResult
from app.models.catalog_models import CatalogAnalyte

# 🌟 IMPORT BỘ NÃO WESTGARD
from app.utils.westgard import eval_rules, get_highest_priority_violation, eval_multilevel

try:
    import pandas as pd
except ImportError:
    pd = None

# ------------------------------------------------------------
# Helpers
# ------------------------------------------------------------
def _to_float(x: Any) -> Optional[float]:
    try:
        if x is None: return None
        s = str(x).strip().replace(",", ".")
        if s == "": return None
        return float(s)
    except Exception:
        return None

def _to_bool(x: Any) -> Optional[int]:
    if x is None: return None
    s = str(x).strip().lower()
    if s in ("1", "true", "yes", "pos", "+", "positive", "dương", "duong"):
        return 1
    if s in ("0", "false", "no", "neg", "-", "negative", "âm", "am"):
        return 0
    return None

# ------------------------------------------------------------
# IQCService (Master Pipeline)
# ------------------------------------------------------------
class IQCService:
    def __init__(self):
        pass

    def __del__(self):
        pass

    def _get_db(self):
        return SessionLocal()

    # -------------- RUN (Phiên chạy) --------------
    def create_run(self, run_date: str, user: str, device: str, department: str,
                   levels_count: int = 2, run_type: str = "quant") -> str:
        run_time = f"{run_date} 12:00:00"
        db = self._get_db()
        try:
            meta_note = json.dumps({
                "levels_count": levels_count,
                "run_type": run_type,
                "device_name": device
            })

            new_run = IQCRun(
                run_date=run_date,
                run_time=run_time,
                department_id=None,
                operator=user,
                device_id=None,
                note=meta_note,
                sync_flag=1
            )
            db.add(new_run)
            db.commit()
            db.refresh(new_run)
            return str(new_run.id)
        except Exception as e:
            db.rollback()
            print(f"❌ [IQCService ERROR] create_run: {e}")
            raise e
        finally:
            db.close()

    def _get_run_meta_orm(self, db, run_id: str) -> Tuple[int, str]:
        run = db.query(IQCRun).filter(IQCRun.id == str(run_id)).first()
        if run and run.note:
            try:
                meta = json.loads(run.note)
                return int(meta.get("levels_count", 2)), str(meta.get("run_type", "quant"))
            except:
                pass
        return 2, "quant"

    # -------------- 🌟 UPSERT RESULTS (WESTGARD PIPELINE) 🌟 --------------
    def upsert_results(self, run_id: str, rows: List[Dict[str, Any]],
                       department: Optional[str] = None,
                       levels_count_default: int = 2,
                       run_type_default: str = "quant") -> int:
        n_saved = 0
        db = self._get_db()

        try:
            levels_count, run_type = self._get_run_meta_orm(db, run_id)
            levels_count = levels_count or levels_count_default
            run_type = run_type or run_type_default

            default_rules = {"1_3s", "1_2s", "2_2s", "R_4s", "4_1s", "10_x", "7_t"}

            for row_data in rows:
                test_code = (row_data.get("test_code") or "").strip()
                if not test_code: continue
                unit = (row_data.get("unit") or "").strip()

                # 🌟 BỘ NHỚ TẠM ĐỂ LƯU KẾT QUẢ CỦA CÁC LEVEL TRƯỚC KHI ĐÁNH GIÁ CHÉO
                test_level_data = {}
                current_z_scores = {}

                for lv_index in range(1, levels_count + 1):
                    level_str = f"L{lv_index}"

                    if level_str not in row_data and f"cat_{level_str}" not in row_data and f"qual_{level_str}" not in row_data:
                        continue

                    val_str, val_num, val_cat, val_bool = None, None, None, None
                    data_found = False

                    if run_type == "quant":
                        val = row_data.get(level_str)
                        vnum = _to_float(val)
                        if vnum is not None:
                            val_str, val_num, data_found = str(val), vnum, True
                    elif run_type == "qual":
                        q_val = row_data.get(f"qual_{level_str}")
                        bv = _to_bool(q_val)
                        if bv is not None:
                            val_str, val_bool, data_found = ("POS" if bv == 1 else "NEG"), bv, True

                    if not data_found: continue

                    # 🌟 BƯỚC 1: Tìm thông số mục tiêu (Mean/SD/Sigma) từ Catalog
                    analyte = db.query(CatalogAnalyte).filter(
                        CatalogAnalyte.test_code == test_code,
                        CatalogAnalyte.level == level_str,
                        CatalogAnalyte.sync_flag != 2
                    ).first()

                    analyte_id = analyte.id if analyte else None
                    lot_id_current = analyte.lot_id if analyte else None

                    mean = analyte.mean if (analyte and analyte.mean is not None) else 0.0
                    sd = analyte.sd if (analyte and analyte.sd) else 0.0

                    # Trích xuất Sigma nếu bạn có lưu trong DB (Nếu không có, trả về None)
                    sigma_val = getattr(analyte, 'sigma', None)

                    z_score = None
                    pass_fail = 1
                    violation_rule = None

                    # 🌟 BƯỚC 2: Kích hoạt Westgard Đơn mức (Single-Level)
                    if run_type == "quant" and sd > 0.0:
                        query_history = db.query(IQCResult.value_num).join(IQCRun).filter(
                            IQCResult.test_code == test_code,
                            IQCResult.level == level_str,
                            IQCResult.value_num.isnot(None)
                        )

                        if lot_id_current:
                            query_history = query_history.join(
                                CatalogAnalyte, IQCResult.analyte_id == CatalogAnalyte.id
                            ).filter(CatalogAnalyte.lot_id == lot_id_current)

                        history_records = query_history.order_by(
                            IQCRun.run_date.desc(),
                            IQCRun.run_time.desc()
                        ).limit(19).all()

                        history_vals = [r[0] for r in reversed(history_records)]
                        history_vals.append(val_num)

                        z_score = round((val_num - mean) / sd, 2)

                        # Truyền thêm Sigma vào não bộ Westgard
                        wg_result = eval_rules(history_vals, mean, sd, default_rules, sigma=sigma_val)

                        if wg_result.get("last_z") is not None:
                            z_score = wg_result.get("last_z")

                        violated = wg_result.get("violated", [])

                        if violated:
                            violation_rule = get_highest_priority_violation(violated)
                            if violation_rule != "1_2s":
                                pass_fail = 0

                    # Lưu tạm vào dict thay vì lưu DB ngay lập tức
                    test_level_data[level_str] = {
                        "val_str": val_str,
                        "val_num": val_num,
                        "analyte_id": analyte_id,
                        "z_score": z_score,
                        "pass_fail": pass_fail,
                        "violation_rule": violation_rule
                    }

                    if z_score is not None:
                        current_z_scores[level_str] = z_score

                # 🌟 BƯỚC 2.5: ĐÁNH GIÁ WESTGARD ĐA MỨC (MULTI-LEVEL ACROSS-RUN)
                if run_type == "quant" and len(current_z_scores) >= 2:
                    multi_violations = eval_multilevel(current_z_scores)

                    for lvl, m_rule in multi_violations.items():
                        s_rule = test_level_data[lvl]["violation_rule"]

                        # So sánh luật đơn và luật đa mức, ưu tiên luật nghiêm trọng hơn
                        combined_rules = [r for r in (s_rule, m_rule) if r]
                        best_rule = get_highest_priority_violation(combined_rules) if combined_rules else None

                        test_level_data[lvl]["violation_rule"] = best_rule

                        # Cập nhật lại Pass/Fail
                        if best_rule and best_rule != "1_2s":
                            test_level_data[lvl]["pass_fail"] = 0

                # 🌟 BƯỚC 3: LƯU THẲNG VÀO ORM SAU KHI ĐÃ ĐÁNH GIÁ CHÉO XONG
                for level_str, data in test_level_data.items():
                    existing_result = db.query(IQCResult).filter(
                        IQCResult.run_id == str(run_id),
                        IQCResult.test_code == test_code,
                        IQCResult.level == level_str
                    ).first()

                    if existing_result:
                        existing_result.value = data["val_str"]
                        existing_result.value_num = data["val_num"]
                        existing_result.unit = unit
                        existing_result.analyte_id = data["analyte_id"]
                        existing_result.z_score = data["z_score"]
                        existing_result.pass_fail = data["pass_fail"]
                        existing_result.violation_rule = data["violation_rule"]
                        existing_result.sync_flag = 1
                    else:
                        new_res = IQCResult(
                            run_id=str(run_id),
                            test_code=test_code,
                            level=level_str,
                            value=data["val_str"],
                            value_num=data["val_num"],
                            unit=unit,
                            analyte_id=data["analyte_id"],
                            z_score=data["z_score"],
                            pass_fail=data["pass_fail"],
                            violation_rule=data["violation_rule"],
                            is_active=1,
                            sync_flag=1
                        )
                        db.add(new_res)
                    n_saved += 1

            db.commit()
            return n_saved
        except Exception as e:
            db.rollback()
            print(f"❌ [IQCService ERROR] upsert_results: {e}")
            raise e
        finally:
            db.close()
    # -------------- LIST/GET RESULTS (Đọc kết quả) --------------

    def get_history(self,
                    department: Optional[str] = None,
                    run_date_from: Optional[str] = None,
                    run_date_to: Optional[str] = None,
                    test_code: Optional[str] = None,
                    lot_no: Optional[str] = None,
                    level: Optional[str] = None,
                    limit: int = 200,
                    sort_order: str = "DESC",
                    active_only: bool = False
                    ) -> List[Dict[str, Any]]:
        """
        Lấy lịch sử kết quả.
        """
        db = self._get_db()
        try:
            # Query cơ bản Join Run và Result
            query = db.query(IQCResult, IQCRun).join(IQCRun, IQCResult.run_id == IQCRun.id)

            # 2. Date Range
            if run_date_from and run_date_to:
                query = query.filter(IQCRun.run_date.between(run_date_from, run_date_to))
            elif run_date_from:
                query = query.filter(IQCRun.run_date == run_date_from)
            elif run_date_to:
                query = query.filter(IQCRun.run_date <= run_date_to)

            # 3. Test Code
            if test_code:
                query = query.filter(func.lower(IQCResult.test_code) == test_code.lower())

            # 4. Level
            if level:
                query = query.filter(func.lower(IQCResult.level) == level.lower())

            # 5. Lot No (Tìm trong Note)
            if lot_no:
                if "%" in lot_no:
                    query = query.filter(IQCResult.note.ilike(f"%{lot_no}%"))
                else:
                    query = query.filter(IQCResult.note.ilike(f"%{lot_no}%"))

            # 6. Active Only
            if active_only:
                query = query.filter(IQCResult.is_active == 1)

            # --- Sắp xếp ---
            if (sort_order or "DESC").upper() == "ASC":
                query = query.order_by(asc(IQCRun.run_date), asc(IQCRun.run_time), asc(IQCResult.id))
            else:
                query = query.order_by(desc(IQCRun.run_date), desc(IQCRun.run_time), desc(IQCResult.id))

            # --- Giới hạn ---
            query = query.limit(limit)

            # --- Execute & Map to Dict ---
            results = query.all()
            data_list = []

            for res, run in results:
                extra = {}
                try:
                    if res.note and "{" in res.note:
                        extra = json.loads(res.note)
                except:
                    pass

                item = {
                    "id": str(res.id),
                    "run_date": run.run_date,
                    "run_time": run.run_time,
                    "user": run.operator,
                    "run_type": extra.get("run_type", "quant"),
                    "test_code": res.test_code,
                    "level": res.level,
                    "department": extra.get("department", ""),
                    "value": res.value,
                    "unit": res.unit,
                    "lot": extra.get("lot", ""),
                    "value_num": res.value_num,
                    "pass_fail": res.pass_fail,
                    "sdi": extra.get("sdi"),
                    "note": res.note,
                    "is_active": res.is_active,
                    "value_cat": extra.get("value_cat"),
                    "value_score": extra.get("value_score"),
                    "value_bool": extra.get("value_bool")
                }
                data_list.append(item)

            return data_list

        except Exception as e:
            print(f"[IQCService ERROR] get_history: {e}")
            return []
        finally:
            db.close()

    # --- ADD NOTE ---
    def add_note_to_result(self, result_id: str, note: str) -> bool:
        db = self._get_db()
        try:
            res = db.query(IQCResult).filter(IQCResult.id == str(result_id)).first()
            if res:
                res.note = note  # Lưu ý: Cân nhắc nối chuỗi nếu muốn giữ JSON cũ
                res.sync_flag = 1
                db.commit()
                return True
            return False
        except Exception as e:
            print(f"[IQCService ERROR] add_note: {e}")
            db.rollback()
            return False
        finally:
            db.close()

    # --- SET ACTIVE ---
    def set_result_active_status(self, result_id: str, is_active: bool) -> bool:
        db = self._get_db()
        try:
            res = db.query(IQCResult).filter(IQCResult.id == str(result_id)).first()
            if res:
                res.is_active = 1 if is_active else 0
                res.sync_flag = 1
                db.commit()
                return True
            return False
        except Exception as e:
            print(f"[IQCService ERROR] set_active: {e}")
            db.rollback()
            return False
        finally:
            db.close()

    # ------------------------------------------------------------
    # IMPORT EXCEL (Đã hoàn thiện logic Mapping)
    # ------------------------------------------------------------
    def import_qc_from_excel(self, file_path: str, lot_id: str) -> Dict[str, Any]:
        """
        Nhập QC từ Excel. Tự động nhận diện Test dựa trên LIMS Code hoặc Test Code.
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            return {"success": False, "msg": "Thiếu thư viện openpyxl. Hãy cài đặt: pip install openpyxl"}

        session = self._get_db()
        count_success = 0

        try:
            # 1. LẤY MAPPING: Tạo từ điển { "TÊN_CỘT_EXCEL" : Analyte_Object }
            # Lấy tất cả xét nghiệm của Lot đang chọn
            analytes = session.query(CatalogAnalyte).filter(
                CatalogAnalyte.lot_id == str(lot_id),
                CatalogAnalyte.sync_flag != 2
            ).all()

            mapping = {}
            for a in analytes:
                # Ưu tiên map theo LIMS Code (nhập từ bảng Mapping Dialog)
                if a.lims_code:
                    mapping[a.lims_code.strip().lower()] = a

                # Fallback: Map theo Test Code (nếu không có LIMS Code)
                if a.test_code:
                    mapping[a.test_code.strip().lower()] = a

                # Fallback: Map theo Tên xét nghiệm
                if a.test_name:
                    mapping[a.test_name.strip().lower()] = a

            if not mapping:
                return {"success": False, "msg": "Lot này chưa có xét nghiệm nào. Hãy vào Danh mục để tạo trước."}

            # 2. ĐỌC FILE EXCEL
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active  # Lấy sheet đầu tiên

            # Lấy dòng Header (Dòng 1)
            headers = []
            for cell in ws[1]:
                headers.append(str(cell.value).strip().lower() if cell.value else "")

            # 3. DUYỆT TỪNG DÒNG DỮ LIỆU (Từ dòng 2 trở đi)
            for row in ws.iter_rows(min_row=2, values_only=True):
                # Cột A (index 0) là Ngày tháng
                raw_date = row[0]
                if not raw_date: continue

                # Xử lý ngày tháng (Hỗ trợ nhiều định dạng)
                run_date_str = dt.datetime.now().strftime("%Y-%m-%d")
                if isinstance(raw_date, (dt.datetime, dt.date)):
                    run_date_str = raw_date.strftime("%Y-%m-%d")
                else:
                    try:
                        # Thử parse chuỗi ngày phổ biến
                        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
                            try:
                                run_date_str = dt.datetime.strptime(str(raw_date), fmt).strftime("%Y-%m-%d")
                                break
                            except:
                                pass
                    except:
                        pass

                # Tạo phiên chạy (Run) mới cho dòng này
                # (User là 'ExcelImport' để dễ phân biệt)
                run_id = self.create_run(
                    run_date=run_date_str,
                    user="ExcelImport",
                    device="Unknown",
                    department="General"
                )

                # Duyệt qua các cột kết quả
                for idx, val in enumerate(row):
                    if idx == 0: continue  # Bỏ qua cột ngày
                    if idx >= len(headers): break

                    header = headers[idx]

                    # KIỂM TRA: Tên cột có trong Mapping không?
                    if header in mapping:
                        analyte = mapping[header]

                        val_num = _to_float(val)  # Hàm helper có sẵn trong file này
                        if val_num is None: continue

                        # Tính toán SDI ngay lập tức
                        # Lưu ý: Dùng getattr để tránh lỗi nếu cột chưa có
                        mean = getattr(analyte, 'mean', 0) or 0
                        sd = getattr(analyte, 'sd', 1) or 1
                        sdi = 0.0
                        if sd > 0:
                            sdi = (val_num - mean) / sd

                        # Lưu metadata vào Note
                        note_data = {
                            "sdi": round(sdi, 2),
                            "lot_id": str(lot_id),
                            "import_source": "excel"
                        }

                        # Tạo kết quả
                        new_res = IQCResult(
                            run_id=str(run_id),
                            test_code=analyte.test_code,
                            # Nếu analyte có level thì dùng, không thì mặc định L1
                            level=analyte.level if hasattr(analyte, 'level') and analyte.level else "L1",
                            value=str(val_num),
                            value_num=val_num,
                            unit=analyte.unit,
                            pass_fail=1,  # Tạm để Đạt, Rule check sẽ chạy sau khi hiển thị
                            is_active=1,
                            sync_flag=1,
                            note=json.dumps(note_data)
                        )
                        session.add(new_res)
                        count_success += 1

            session.commit()
            return {"success": True, "msg": f"Đã nhập thành công {count_success} kết quả."}

        except Exception as e:
            session.rollback()
            return {"success": False, "msg": f"Lỗi Import: {str(e)}"}
        finally:
            session.close()

    # =========================================================================
    # [CORE] LOGIC GIẢI MÃ DỮ LIỆU (Dùng chung cho FILE và TCP)
    # =========================================================================
    def _parse_raw_data(self, content: str) -> List[Dict]:
        """Hàm nội bộ: Chuyển đổi chuỗi thô (ASTM/Text) thành danh sách kết quả"""
        results = []
        try:
            # 1. Nhận diện ASTM (Có chứa các ký tự đặc trưng như R|... hoặc H|\...)
            if "R|" in content or "H|" in content:
                lines = content.split('\n')
                for line in lines:
                    line = line.strip()
                    if line.startswith('R|'):  # Dòng Result trong ASTM
                        parts = line.split('|')
                        # Cấu trúc ASTM chuẩn: R|seq|TEST^^^|...|VALUE|...
                        # Index thường là: 2=TestCode, 8=Value (tùy máy)
                        if len(parts) >= 9:
                            test_code = parts[2].strip()
                            value = parts[8].strip()

                            # Xử lý mã phụ (VD: GOT^...)
                            if '^' in test_code:
                                test_code = test_code.split('^')[0]

                            # Một số máy đặt Value ở index khác, cần linh hoạt check
                            if not value and len(parts) > 3:
                                # Fallback logic nếu cần
                                pass

                            if test_code and value:
                                results.append({"test_code": test_code, "value": value})

            # 2. Nhận diện Text/CSV đơn giản (Test,Value)
            else:
                lines = content.split('\n')
                for line in lines:
                    line = line.strip()
                    if not line: continue

                    parts = []
                    if ',' in line:
                        parts = line.split(',')
                    elif '\t' in line:
                        parts = line.split('\t')
                    elif '|' in line:
                        parts = line.split('|')
                    else:
                        parts = line.split()

                    if len(parts) >= 2:
                        results.append({
                            "test_code": parts[0].strip(),
                            "value": parts[1].strip()
                        })
        except Exception as e:
            print(f"[Parser] Lỗi giải mã: {e}")

        return results

    # =========================================================================
    # HÀM DÙNG CHUNG: PHÂN TÍCH DỮ LIỆU (CORE PARSER)
    # =========================================================================
    def _parse_data_common(self, content: str) -> List[Dict]:
        """
        [CORE] Phân tích chuỗi dữ liệu (ASTM hoặc Text/CSV).
        - Tự động tìm Giá trị ở index 8 (Sphera), 3, hoặc 4.
        - Tự động lấy Đơn vị (Unit) ở index 4.
        """
        results = []
        try:
            lines = content.splitlines()
            for line in lines:
                line = line.strip()
                if not line: continue

                # --- A. XỬ LÝ ASTM (Bắt đầu bằng R|) ---
                if 'R|' in line:
                    parts = line.split('|')
                    # Cấu trúc chuẩn: R | Seq | Test | Result | Unit | ... | ... | Value(Sphera)
                    # Index:          0   1     2      3        4      ...   ...   8

                    if len(parts) > 2:
                        test_code = parts[2].strip()
                        value = ""
                        unit = ""

                        # 1. LOGIC TÌM GIÁ TRỊ (VALUE)
                        # Ưu tiên 1: Index 8 (Máy Sphera)
                        if len(parts) > 8 and parts[8].strip():
                            value = parts[8].strip()
                        # Ưu tiên 2: Index 3 (ASTM Chuẩn - Result)
                        elif len(parts) > 3 and parts[3].strip():
                            # Check kỹ nếu là số
                            val_check = parts[3].strip()
                            if val_check.replace('.', '', 1).isdigit(): value = val_check
                        # Ưu tiên 3: Index 4 (Hiếm gặp, nhưng một số máy để Value ở đây)
                        elif len(parts) > 4 and parts[4].strip():
                            val_check = parts[4].strip()
                            # Chỉ lấy nếu nó là số thuần túy
                            if val_check.replace('.', '', 1).isdigit(): value = val_check

                        # 2. LOGIC TÌM ĐƠN VỊ (UNIT)
                        # Thường nằm ở index 4.
                        # Chỉ lấy nếu index 4 KHÔNG PHẢI là Value đã lấy ở trên.
                        if len(parts) > 4:
                            candidate = parts[4].strip()
                            if candidate and candidate != value:
                                unit = candidate

                        # Xử lý Test Code (Bỏ mã phụ sau ^, VD: GOT^...)
                        if '^' in test_code:
                            test_code = test_code.split('^')[0]

                        if test_code and value:
                            results.append({
                                "test_code": test_code,
                                "value": value,
                                "unit": unit  # Trả về thêm Unit
                            })

                # --- B. XỬ LÝ TEXT/CSV ---
                elif ',' in line or '\t' in line:
                    sep = ',' if ',' in line else '\t'
                    parts = line.split(sep)
                    if len(parts) >= 2:
                        t_c = parts[0].strip()
                        t_v = parts[1].strip()
                        # Cố gắng lấy Unit ở cột 3 nếu có
                        t_u = parts[2].strip() if len(parts) > 2 else ""

                        # Lọc dòng tiêu đề rác
                        if t_v.lower() not in ['value', 'kq', 'result', 'unit']:
                            results.append({
                                "test_code": t_c,
                                "value": t_v,
                                "unit": t_u
                            })

        except Exception as e:
            print(f"[Parser] Lỗi phân tích: {e}")

        return results

    # =========================================================================
    # 5. DEVICE INTEGRATION (FILE - Ưu tiên tên File = Tên Lot)
    # =========================================================================
    def read_machine_file_by_lot(self, folder_path: str, lot_code: str, target_date: dt.date) -> List[Dict]:
        """
        Đọc file kết quả máy (Tìm theo tên Lot & Ngày).
        Logic:
        1. Quét thư mục tìm file có tên = lot_code (VD: 1234.astm).
        2. Nếu thấy, đọc nội dung và phân tích.
        3. [MỚI] Tự động gán mã Lot từ tên file vào kết quả (để đảm bảo không bị sót nếu nội dung file thiếu header Lot).
        """
        # 1. Validate đầu vào
        if not folder_path or not os.path.exists(folder_path) or not lot_code:
            return []

        # 2. Tìm file (Quét thư mục - Không phân biệt hoa thường)
        # Cách này giúp ta CHỈ MỞ 1 FILE duy nhất, không cần mở từng file để check.
        target_file = None
        SUPPORTED_EXTS = ['.astm', '.txt', '.csv']

        try:
            all_files = os.listdir(folder_path)
            for fname in all_files:
                name, ext = os.path.splitext(fname)
                # So sánh tên file với Lot Code (case-insensitive)
                if name.lower() == lot_code.lower() and ext.lower() in SUPPORTED_EXTS:
                    full_path = os.path.join(folder_path, fname)

                    # Check ngày sửa đổi
                    mtime = os.path.getmtime(full_path)
                    file_date = dt.datetime.fromtimestamp(mtime).date()
                    if file_date == target_date:
                        target_file = full_path
                        break
                    # else:
                    #    print(f"[IQC] Thấy file {fname} nhưng sai ngày")
        except Exception as e:
            print(f"[IQC] Lỗi quét thư mục: {e}")
            return []

        if not target_file:
            return []

        # 3. Đọc nội dung & Gọi hàm Parse chung
        encodings_to_try = ['utf-8', 'latin-1', 'cp1252']
        results = []

        content = ""
        read_success = False

        for enc in encodings_to_try:
            try:
                with open(target_file, 'r', encoding=enc) as f:
                    content = f.read()
                    read_success = True
                    break
            except UnicodeDecodeError:
                continue
            except Exception:
                break

        if read_success:
            results = self._parse_data_common(content)

            # [CẬP NHẬT QUAN TRỌNG]
            # Bổ sung tên Lot vào kết quả dựa trên tên File.
            # Vì ta đã tìm đúng file "1234.astm", nên chắc chắn kết quả trong đó thuộc về Lot "1234".
            for item in results:
                # Nếu parser chưa tìm thấy lot trong nội dung (hoặc file text đơn giản),
                # thì gán cứng bằng tên file (lot_code).
                if not item.get('lot'):
                    item['lot'] = lot_code

        return results

    # =========================================================================
    # 6. TCP/IP INTEGRATION (LAN - Dựa trên logic của bạn)
    # =========================================================================
    def receive_data_via_tcp(self, ip: str, port: int, timeout: int = 5) -> List[Dict]:
        """
        Kết nối TCP/IP để nhận dữ liệu.
        """
        raw_data = b""
        import socket
        import time

        print(f"[TCP] Đang kết nối tới {ip}:{port}...")

        try:
            with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
                s.settimeout(timeout)
                s.connect((ip, int(port)))

                # Chờ server kịp gửi
                time.sleep(0.5)

                start_time = time.time()
                while True:
                    try:
                        chunk = s.recv(4096)
                        if not chunk: break  # Kết nối đóng

                        raw_data += chunk

                        # Check ký tự kết thúc ASTM
                        if b'\x04' in chunk or b'\x03' in chunk:
                            print("[TCP] Đã nhận ký tự kết thúc (EOT/ETX).")
                            break

                        # Logic chờ thêm gói tin sót
                        if raw_data and len(chunk) < 4096:
                            s.settimeout(1.0)

                    except socket.timeout:
                        if raw_data:
                            print("[TCP] Timeout (đã có dữ liệu).")
                            break
                        if time.time() - start_time > timeout:
                            print("[TCP] Timeout (chưa có dữ liệu).")
                            break
                        continue

        except Exception as e:
            print(f"[TCP] Lỗi kết nối: {e}")
            return []

        if not raw_data:
            print("[TCP] Buffer trống.")
            return []

        print(f"[TCP] Nhận {len(raw_data)} bytes.")

        # Decode & Gọi hàm Parse chung
        content = ""
        encodings = ['utf-8', 'latin-1', 'cp1252']
        for enc in encodings:
            try:
                content = raw_data.decode(enc)
                break
            except:
                continue

        # [QUAN TRỌNG] Gọi hàm chung, không viết lại logic parse ở đây
        results = self._parse_data_common(content)

        print(f"[TCP] Phân tích được: {len(results)} kết quả.")
        return results

    # =========================================================================
    # HÀM DÙNG CHUNG: PHÂN TÍCH DỮ LIỆU (PARSE)
    # =========================================================================
    def _parse_data_common(self, content: str) -> List[Dict]:
        """
        Hàm nội bộ: Phân tích chuỗi ASTM/Text -> Danh sách kết quả + Đơn vị
        Dùng chung cho cả đọc File và TCP để đảm bảo logic đồng nhất.
        """
        results = []
        try:
            lines = content.splitlines()
            for line in lines:
                line = line.strip()
                if not line: continue

                # --- TRƯỜNG HỢP 1: ASTM (Bắt đầu bằng R|) ---
                if 'R|' in line:
                    parts = line.split('|')
                    # Cấu trúc ASTM: R | Seq | Test | Result | Unit | ... | ... | Value(Sphera)
                    # Index:         0   1     2      3        4      ...   ...   8

                    if len(parts) > 2:
                        test_code = parts[2].strip()
                        value = ""
                        unit = ""

                        # 1. Lấy GIÁ TRỊ (Ưu tiên index 8, rồi đến 3, 4)
                        if len(parts) > 8 and parts[8].strip():
                            value = parts[8].strip()
                        elif len(parts) > 3 and parts[3].strip():
                            # Check kỹ nếu index 3 là số
                            val_check = parts[3].strip()
                            if val_check.replace('.', '', 1).isdigit(): value = val_check
                        elif len(parts) > 4 and parts[4].strip():
                            # Một số máy để value ở index 4 (hiếm gặp nhưng có)
                            if parts[4].strip().replace('.', '', 1).isdigit(): value = parts[4].strip()

                        # 2. [QUAN TRỌNG] Lấy ĐƠN VỊ (Thường ở index 4)
                        # Lưu ý: Nếu index 4 đã bị lấy làm Value ở trên thì thôi, còn không thì nó là Unit
                        if len(parts) > 4:
                            candidate_unit = parts[4].strip()
                            # Nếu candidate_unit không phải là số (Value) thì nó là Unit
                            if not candidate_unit.replace('.', '', 1).isdigit():
                                unit = candidate_unit

                        # Xử lý Test Code (Bỏ mã phụ sau ^)
                        if '^' in test_code: test_code = test_code.split('^')[0]

                        if test_code and value:
                            results.append({
                                "test_code": test_code,
                                "value": value,
                                "unit": unit  # <-- Trả về Unit
                            })

                # --- TRƯỜNG HỢP 2: TEXT/CSV (Phẩy hoặc Tab) ---
                elif ',' in line or '\t' in line:
                    sep = ',' if ',' in line else '\t'
                    parts = line.split(sep)
                    if len(parts) >= 2:
                        t_c = parts[0].strip()
                        t_v = parts[1].strip()

                        # Nếu file text có cột 3 là Unit (VD: GLU, 5.5, mmol/L)
                        t_u = parts[2].strip() if len(parts) > 2 else ""

                        if t_v.lower() not in ['value', 'kq', 'result']:
                            results.append({"test_code": t_c, "value": t_v, "unit": t_u})

        except Exception as e:
            print(f"[Parser] Lỗi: {e}")

        return results

    # =========================================================================
    # 7. RS232 INTEGRATION (SERIAL PORT)
    # =========================================================================
    def receive_data_via_serial(self, port: str, baudrate: int,
                                parity: str = 'N', stopbits: int = 1,
                                bytesize: int = 8, timeout: int = 5) -> List[Dict]:
        """
        Kết nối qua cổng COM (RS232) để nhận dữ liệu.
        """
        raw_data = b""
        import time
        try:
            import serial  # Yêu cầu thư viện pyserial
        except ImportError:
            print("[Serial] Chưa cài đặt module 'pyserial'")
            return []

        print(f"[Serial] Đang mở cổng {port} (Baud: {baudrate})...")

        ser = None
        try:
            # Map tham số Parity từ UI/DB sang Pyserial constant
            parity_map = {'N': serial.PARITY_NONE, 'E': serial.PARITY_EVEN, 'O': serial.PARITY_ODD}
            p_val = parity_map.get(str(parity).upper(), serial.PARITY_NONE)

            # Khởi tạo kết nối Serial
            ser = serial.Serial(
                port=port,
                baudrate=int(baudrate),
                bytesize=int(bytesize),
                parity=p_val,
                stopbits=int(stopbits),
                timeout=timeout
            )

            # Chờ một chút để cổng ổn định
            time.sleep(0.5)

            # Vòng lặp đọc dữ liệu
            start_time = time.time()
            while True:
                # Đọc từng block (nếu có)
                if ser.in_waiting > 0:
                    chunk = ser.read(ser.in_waiting)
                    raw_data += chunk

                    # [ASTM] Kiểm tra ký tự kết thúc (EOT \x04 hoặc ETX \x03)
                    # File LOT1.astm của bạn kết thúc bằng L||N (thường đi kèm ETX/EOT ẩn)
                    if b'\x04' in chunk or b'\x03' in chunk:
                        print("[Serial] Đã nhận ký tự kết thúc.")
                        break

                    # Reset lại timeout cục bộ nếu đang nhận dữ liệu
                    start_time = time.time()
                else:
                    # Nếu buffer rỗng
                    # 1. Nếu đã có dữ liệu rồi mà bỗng nhiên ngắt -> Đợi thêm 1s rồi dùng
                    if raw_data and (time.time() - start_time > 1.0):
                        print("[Serial] Dữ liệu đã ngừng gửi.")
                        break

                    # 2. Nếu chưa có dữ liệu -> Chờ đến khi hết Timeout tổng (5s)
                    if time.time() - start_time > timeout:
                        print("[Serial] Timeout (Không có dữ liệu).")
                        break

                    time.sleep(0.1)  # Ngủ ngắn để đỡ ngốn CPU

        except Exception as e:
            print(f"[Serial] Lỗi kết nối: {e}")
            return []
        finally:
            if ser and ser.is_open:
                ser.close()
                print("[Serial] Đã đóng cổng.")

        if not raw_data:
            print("[Serial] Buffer trống.")
            return []

        print(f"[Serial] Nhận {len(raw_data)} bytes.")

        # Decode
        content = ""
        encodings = ['utf-8', 'latin-1', 'cp1252']
        for enc in encodings:
            try:
                content = raw_data.decode(enc)
                break
            except:
                continue

        # [QUAN TRỌNG] Gọi lại hàm parse chung (để lấy Value index 8 và Unit index 4)
        results = self._parse_data_common(content)

        print(f"[Serial] Phân tích được: {len(results)} kết quả.")
        return results

    # [THÊM VÀO FILE iqc_service.py]

    def delete_result(self, result_id: str) -> bool:
        """
        Xóa vĩnh viễn một kết quả IQC khỏi Database (Hard Delete).
        """
        db = self._get_db()
        try:
            # Tìm và xóa dòng dữ liệu có id tương ứng
            db.query(IQCResult).filter(IQCResult.id == str(result_id)).delete()
            db.commit()
            return True
        except Exception as e:
            db.rollback()
            print(f"[IQCService] Delete Result Error: {e}")
            return False
        finally:
            db.close()

# [THÊM VÀO CUỐI FILE iqc_service.py, TRONG CLASS IQCService]

    def update_result_value(self, result_id: str, new_val: str, new_unit: str, run_type: str = "quant") -> bool:
        """
        Cập nhật giá trị kết quả (Sửa từ Admin).
        Tự động cập nhật value_num, value_bool tùy theo loại xét nghiệm.
        """
        db = self._get_db()
        try:
            res = db.query(IQCResult).filter(IQCResult.id == str(result_id)).first()
            if not res:
                return False

            # Cập nhật thông tin chung
            res.value = new_val
            res.unit = new_unit
            res.sync_flag = 1  # Đánh dấu để đồng bộ nếu cần

            # Cập nhật thông tin chi tiết theo loại
            rt = str(run_type).lower()
            if rt == 'quant':
                res.value_num = _to_float(new_val) # Hàm _to_float đã có sẵn ở đầu file
            elif rt == 'qual':
                b_val = _to_bool(new_val) # Hàm _to_bool đã có sẵn ở đầu file
                res.value_bool = b_val
                # Chuẩn hóa text hiển thị cho định tính
                if b_val == 1: res.value = "POS"
                elif b_val == 0: res.value = "NEG"
            elif rt == 'semi':
                res.value_cat = new_val
                # Nếu nhập số cho bán định lượng, cũng lưu vào value_score
                res.value_score = _to_float(new_val)

            db.commit()
            return True
        except Exception as e:
            db.rollback()
            print(f"[Update Error] {e}")
            return False
        finally:
            db.close()